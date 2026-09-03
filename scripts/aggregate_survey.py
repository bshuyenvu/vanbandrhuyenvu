"""Aggregate a Google Forms-style survey .xlsx into stats + comments.

Pattern: file Excel kết quả Google Forms với các cột:
  - Vai trò người trả lời (categorical, vd: "Giáo viên/Nhân viên", "Ban Giám hiệu/CBQL")
  - Nhiều cột thang điểm 1-5 (numeric)
  - Vài cột góp ý text (open-ended)

Usage:
    python aggregate_survey.py <file.xlsx> [--out <stats.yaml>]
    python aggregate_survey.py <file.xlsx> --comments-only

Output: 3 phần
  1. Demographics: số phản hồi theo vai trò
  2. Stats: trung bình + min/max + count cho mỗi cột số
  3. Comments: ý kiến text non-trivial (loại bỏ "không", "không có ý kiến", "k", v.v.)
"""
import argparse
import io
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

TRIVIAL_PATTERNS = [
    r"^không$", r"^không có ý kiến$", r"^không có$", r"^không có đè xuất$",
    r"^không có đề xuất$", r"^k$", r"^kh$", r"^o$", r"^/$", r"^-+$",
    r"^không có gì$", r"^nothing$", r"^no$", r"^none$", r"^x$", r"^\.$",
    r"^chưa có$", r"^chưa$", r"^đồng ý$", r"^nhất trí$", r"^nhat tri$",
    r"^không\s+(có\s+)?(ý\s+kiến|gì|đề\s+xuất|bổ\s+sung).*$",
]
TRIVIAL_RE = re.compile("|".join(TRIVIAL_PATTERNS), re.IGNORECASE)


def is_trivial_comment(text):
    if not text or not str(text).strip():
        return True
    t = str(text).strip()
    if len(t) < 4:
        return True
    if TRIVIAL_RE.fullmatch(t):
        return True
    # Strip whitespace and punctuation, re-check
    t_clean = re.sub(r"[\s\.,;:!?]+", "", t.lower())
    if t_clean in {"khong", "khongco", "khongcoykien", "khongcodexuat", "k", "no"}:
        return True
    return False


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", default=None, help="Optional path to save stats as YAML")
    ap.add_argument("--comments-only", action="store_true")
    ap.add_argument("--role-col", type=int, default=None,
                    help="1-indexed column number for respondent role (default: auto-detect)")
    args = ap.parse_args()

    path = Path(args.xlsx).resolve()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    # Read headers
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v else "")

    # Detect role column
    role_col_idx = args.role_col
    if role_col_idx is None:
        for i, h in enumerate(headers, 1):
            if "vai trò" in h.lower() or "role" in h.lower() or "đối tượng" in h.lower():
                role_col_idx = i
                break

    # Read rows
    rows = []
    for r in range(2, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(row=r, column=c).value)
        # Skip if entire row is empty
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rows.append(row)

    n = len(rows)
    print(f"=== Aggregating {path.name} ===")
    print(f"Total responses: {n}\n")

    # 1) Demographics by role
    if role_col_idx:
        print(f"--- DEMOGRAPHICS (col {role_col_idx}: {headers[role_col_idx-1][:80]}) ---")
        roles = Counter(str(r[role_col_idx-1]).strip() for r in rows if r[role_col_idx-1])
        for role, cnt in roles.most_common():
            print(f"  {cnt:4d}  ({100*cnt/n:.1f}%)  {role}")
        print()

    # 2) Numeric stats per column
    if not args.comments_only:
        print("--- STATS PER COLUMN (numeric only) ---")
        for ci, h in enumerate(headers, 1):
            vals = [safe_float(r[ci-1]) for r in rows]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            avg = sum(vals) / len(vals)
            mn, mx = min(vals), max(vals)
            # Distribution for 1-5 scale
            dist = Counter(int(v) for v in vals if 1 <= v <= 5)
            dist_str = " ".join(f"{k}:{dist[k]}" for k in sorted(dist))
            short = h[:90].replace("\n", " ")
            print(f"\n  [C{ci}] {short}")
            print(f"    n={len(vals)} avg={avg:.2f} min={mn} max={mx}  dist[{dist_str}]")
        print()

    # 3) Comments
    print("--- COMMENTS (non-trivial) ---")
    COMMENT_KEYWORDS = ["góp ý", "đề xuất", "ý kiến", "bình luận", "comment",
                         "feedback", "kiến nghị", "khuyến nghị", "nhận xét"]
    text_cols = []
    for ci, h in enumerate(headers, 1):
        h_lower = h.lower()
        vals = [str(r[ci-1]) for r in rows if r[ci-1] is not None and isinstance(r[ci-1], str)]
        if not vals or len(vals) < 5:
            continue
        # Skip categorical: < 5 unique values among non-empty (likely a dropdown)
        unique_count = len(set(v.strip().lower() for v in vals))
        if unique_count < 5:
            continue
        # Strong signal: header contains a comment keyword
        if any(kw in h_lower for kw in COMMENT_KEYWORDS):
            text_cols.append(ci)
            continue
        # Weak signal: avg cell length > 50 chars (long-form text)
        avg_len = sum(len(v) for v in vals) / len(vals)
        if avg_len > 50:
            text_cols.append(ci)

    for ci in text_cols:
        h = headers[ci-1]
        comments = []
        for ri, r in enumerate(rows, 2):
            v = r[ci-1]
            if v is None:
                continue
            t = str(v).strip()
            if is_trivial_comment(t):
                continue
            # Add role prefix if available
            role = str(r[role_col_idx-1]).strip() if role_col_idx else ""
            comments.append((ri, role, t))

        short_h = h[:90].replace("\n", " ")
        print(f"\n  >> [C{ci}] {short_h}")
        print(f"     {len(comments)} non-trivial / {len(rows)} total\n")
        for ri, role, t in comments[:30]:  # cap at 30 per column
            tprev = t[:300].replace("\n", " | ")
            print(f"     - (row {ri}, {role[:30]}) {tprev}")
        if len(comments) > 30:
            print(f"     ... +{len(comments) - 30} comments more")

    # Save YAML if requested
    if args.out:
        out_path = Path(args.out).resolve()
        # Build a simple YAML manually (no PyYAML dependency)
        lines = []
        lines.append(f"file: {path.name}")
        lines.append(f"total_responses: {n}")
        if role_col_idx:
            lines.append("demographics:")
            for role, cnt in roles.most_common():
                role_safe = role.replace('"', '\\"')
                lines.append(f'  - role: "{role_safe}"')
                lines.append(f"    count: {cnt}")
                lines.append(f"    percent: {100*cnt/n:.1f}")

        lines.append("stats:")
        for ci, h in enumerate(headers, 1):
            vals = [safe_float(r[ci-1]) for r in rows]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            avg = sum(vals) / len(vals)
            h_safe = h.replace('"', '\\"').replace("\n", " ")
            lines.append(f'  - col: {ci}')
            lines.append(f'    header: "{h_safe[:200]}"')
            lines.append(f'    n: {len(vals)}')
            lines.append(f'    avg: {avg:.3f}')
            lines.append(f'    min: {min(vals)}')
            lines.append(f'    max: {max(vals)}')

        lines.append("comments:")
        for ci in text_cols:
            h = headers[ci-1]
            h_safe = h.replace('"', '\\"').replace("\n", " ")
            comments = []
            for r in rows:
                v = r[ci-1]
                if v is None:
                    continue
                t = str(v).strip()
                if not is_trivial_comment(t):
                    comments.append(t)
            lines.append(f'  - col: {ci}')
            lines.append(f'    header: "{h_safe[:200]}"')
            lines.append(f'    non_trivial_count: {len(comments)}')
            lines.append(f'    items:')
            for c in comments:
                c_safe = c.replace('"', '\\"').replace("\n", " ")[:500]
                lines.append(f'      - "{c_safe}"')

        out_path.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n✓ Saved YAML: {out_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
